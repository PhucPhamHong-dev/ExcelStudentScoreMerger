import os
from flask import Flask, request, render_template, send_file
from werkzeug.utils import secure_filename
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'replace-with-your-secret'


def is_processed_excel(file_stream):
    """
    Check for our hidden marker sheet '_processed_marker'.
    """
    file_stream.seek(0)
    try:
        wb = load_workbook(filename=BytesIO(file_stream.read()), read_only=True)
        return '_processed_marker' in wb.sheetnames
    except Exception:
        return False


def merge_scores(raw_df: pd.DataFrame, template_df: pd.DataFrame) -> pd.DataFrame:
    # Ensure required columns in raw
    if 'Login' not in raw_df.columns or 'Mark(10)' not in raw_df.columns:
        raise KeyError("Raw file missing 'Login' or 'Mark(10)' column")
    lookup = dict(zip(raw_df['Login'], raw_df['Mark(10)']))
    # Ensure required columns in template
    for col in ['Class', 'RollNumber', 'Mark', 'Note']:
        if col not in template_df.columns:
            raise KeyError(f"Template file missing '{col}' column")
    # Map marks
    template_df['Mark'] = template_df['RollNumber'].map(lookup)
    return template_df


@app.route("/", methods=["GET", "POST"])
def index():
    error = None
    if request.method == "POST":
        raw_file = request.files.get("raw_exam")
        template_file = request.files.get("template")
        # 1) kiểm tra upload
        if not raw_file or not template_file:
            error = "Vui lòng chọn cả hai file."
            return render_template("index.html", error=error)

  
        def has_processed_suffix(fname):
            base, _ = os.path.splitext(fname.lower())
            return base.endswith(("_filled","_merged"))
        if has_processed_suffix(raw_file.filename) or has_processed_suffix(template_file.filename):
            error = "File dường như đã được xử lý rồi. Vui lòng chọn file gốc."
            return render_template("index.html", error=error)


        if is_processed_excel(raw_file.stream) or is_processed_excel(template_file.stream):
            error = "File đã qua xử lý (có marker). Vui lòng chọn file gốc."
            return render_template("index.html", error=error)

  
        try:
            xls = pd.ExcelFile(raw_file, engine='openpyxl')
            raw_df = None
            for sheet in xls.sheet_names:
                df = xls.parse(sheet)
                if {'Login', 'Mark(10)'}.issubset(df.columns):
                    raw_df = df
                    break
            if raw_df is None:
                error = "File RAW thiếu cột bắt buộc 'Login' và 'Mark(10)' trong tất cả các sheet."
                return render_template("index.html", error=error)
        except Exception as e:
            error = f"Lỗi khi đọc file RAW: {e}"
            return render_template("index.html", error=error)


        try:
            xls2 = pd.ExcelFile(template_file, engine='openpyxl')
            template_df = xls2.parse(xls2.sheet_names[0])
        except Exception as e:
            error = f"Lỗi khi đọc file Template: {e}"
            return render_template("index.html", error=error)


        try:
            merged_df = merge_scores(raw_df, template_df)
        except KeyError as e:
            error = str(e)
            return render_template("index.html", error=error)
        except Exception as e:
            error = f"Lỗi khi merge điểm: {e}"
            return render_template("index.html", error=error)


        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            merged_df.to_excel(writer, index=False, sheet_name='Sheet1')
            wb = writer.book
            ws = wb.create_sheet('_processed_marker')
            ws.sheet_state = 'veryHidden'
        # ExcelWriter context manager saves automatically
        output.seek(0)

        base, ext = os.path.splitext(secure_filename(template_file.filename))
        output_name = f"{base}_filled{ext}"
        return send_file(
            output,
            as_attachment=True,
            download_name=output_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    return render_template("index.html", error=error)


if __name__ == "__main__":
    app.run(debug=True)